# core/mps.py
# MPS mensal (não semanal) com opção de:
# - Estoque de segurança variável por mês (safety_stock_series)
# - Congelamento de horizonte (frozen_range)
# Mantém retrocompatibilidade com chamadas antigas.

from __future__ import annotations
import math, re
from pathlib import Path
from typing import Iterable, Optional, Dict, List, Tuple, Union
import pandas as pd
import numpy as np

# ====== CAMINHO DE SAÍDA FIXO (ajuste se necessário) ======
OUTPUT_DIR = Path(r"C:\Users\vitor\OneDrive\TCC\Códigos VSCODE\MRP e MPS")

# ======================= Helpers de período mensal =======================

_PT_MONTHS = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out":10, "nov":11, "dez":12
}
_PT_MONTHS_REV = {v: k.capitalize() for k, v in _PT_MONTHS.items()}

def _to_period_m(value: Union[str, pd.Timestamp, pd.Period]) -> pd.Period:
    """Converte 'Set/25', '2025-09', datetime etc. em Period('YYYY-MM','M')."""
    if isinstance(value, pd.Period):
        return value.asfreq('M')
    if isinstance(value, pd.Timestamp):
        return value.to_period('M')
    s = str(value).strip()
    m = re.match(r"^([A-Za-zçÇ]{3})[\/\-\s](\d{2,4})$", s, flags=re.IGNORECASE)
    if m:
        mon_txt = m.group(1).lower()[:3]
        year_txt = m.group(2)
        if mon_txt not in _PT_MONTHS:
            raise ValueError(f"Mês PT inválido: {mon_txt}")
        year = int(year_txt)
        if year < 100:
            year += 2000  # pivot simples (25 -> 2025)
        month = _PT_MONTHS[mon_txt]
        return pd.Period(freq='M', year=year, month=month)
    m2 = re.match(r"^(\d{4})[\/\-](\d{1,2})$", s)
    if m2:
        year = int(m2.group(1)); month = int(m2.group(2))
        return pd.Period(freq='M', year=year, month=month)
    try:
        return pd.Period(pd.to_datetime(s), 'M')
    except Exception as e:
        raise ValueError(f"Não consegui interpretar o mês: {value}") from e

def _label_pt(period: pd.Period) -> str:
    """Formata Period('2025-09','M') => 'Set/25'."""
    mon = _PT_MONTHS_REV[period.month]
    yy = str(period.year)[-2:]
    return f"{mon}/{yy}"

# ======================= Cálculo do MPS (mensal) =======================

def compute_mps_monthly(
    df_forecast: pd.DataFrame,            # colunas: ds, y
    *,
    lot_policy: str = "FX",               # "L4L" ou "FX"
    lot_size: int = 100,
    safety_stock: int = 0,
    lead_time: int = 1,                   # em meses
    initial_inventory: int = 0,
    scheduled_receipts: Optional[Dict[Union[str, pd.Period], int]] = None,  # {"Set/25": qtd, ...}
    firm_customer_orders: Optional[pd.DataFrame] = None,  # colunas: ds, y  (opcional)

    # -------- NOVOS PARÂMETROS OPCIONAIS (retrocompatíveis) --------
    safety_stock_series: Optional[Iterable[int]] = None,       # SS por mês (mesmo T do horizonte)
    frozen_range: Optional[Tuple[Union[str, pd.Period], Union[str, pd.Period]]] = None,  # intervalo inclusivo
) -> pd.DataFrame:
    """
    Retorna DataFrame com:
      period (Period[M]), period_label, gross_requirements, scheduled_receipts,
      on_hand_begin, net_requirements, planned_order_receipts,
      planned_order_releases, projected_on_hand_end, atp (se houver pedidos).

    NOVO:
      - safety_stock_series: se fornecido, aplica SS específico por mês (senão usa safety_stock fixo).
      - frozen_range: intervalo (inclusive) de períodos 'congelados' (não planeja recebimentos nesses buckets).
        Política adotada: CONGELAR RECEBIMENTOS no intervalo (planned_order_receipts = 0). Recebimentos agendados (scheduled_receipts) continuam válidos.
        Efeito: se faltar estoque dentro do fence, a compensação acontece fora dele.
    """
    if not {"ds", "y"}.issubset(df_forecast.columns):
        raise ValueError("df_forecast deve ter colunas: ds, y")

    # Ordena e normaliza períodos
    periods = pd.Series(df_forecast["ds"]).map(_to_period_m)
    y_fore = df_forecast["y"].astype(float).round().astype(int).tolist()
    order = periods.argsort().to_list()
    periods = periods.iloc[order].reset_index(drop=True)
    y_fore = [y_fore[i] for i in order]
    T = len(periods)

    # scheduled receipts -> mapa por idx
    sr_in = scheduled_receipts or {}
    sr_map = { _to_period_m(k): int(v) for k, v in sr_in.items() }
    sr_idx: Dict[int, int] = { i: sr_map.get(periods[i], 0) for i in range(T) }

    # pedidos firmes
    cust = None
    if firm_customer_orders is not None:
        if not {"ds", "y"}.issubset(firm_customer_orders.columns):
            raise ValueError("firm_customer_orders deve ter colunas: ds, y")
        cust_series = (firm_customer_orders
                       .assign(period=lambda d: d["ds"].map(_to_period_m))
                       .groupby("period")["y"].sum())
        cust = [int(round(cust_series.get(p, 0))) for p in periods]

    # validação parâmetros básicos
    if lot_policy.upper() not in ("L4L", "FX"):
        raise ValueError("lot_policy deve ser 'L4L' ou 'FX'")
    lot_size = max(1, int(lot_size))
    safety_stock = max(0, int(safety_stock))
    lead_time = max(0, int(lead_time))
    on_hand = int(initial_inventory)

    # ---------- SS por mês (se fornecido) ----------
    ss_by_idx: Optional[List[int]] = None
    if safety_stock_series is not None:
        arr = list(map(lambda x: max(0, int(x)), safety_stock_series))
        if len(arr) != T:
            raise ValueError(f"safety_stock_series deve ter comprimento {T} (recebi {len(arr)})")
        ss_by_idx = arr

    # ---------- Índices congelados (se fornecido) ----------
    frozen_idx: set[int] = set()
    if frozen_range is not None:
        f0 = _to_period_m(frozen_range[0])
        f1 = _to_period_m(frozen_range[1])
        fmin, fmax = (f0, f1) if f0 <= f1 else (f1, f0)
        for i, p in enumerate(periods):
            if fmin <= p <= fmax:
                frozen_idx.add(i)

    rows = []
    por = [0]*T   # planned_order_receipts por período i
    pol = [0]*T   # planned_order_releases por período i (após deslocamento LT)

    for i in range(T):
        gross = int(y_fore[i])
        rec_sched = int(sr_idx.get(i, 0))
        on_hand_begin = on_hand + rec_sched

        # SS efetivo do mês (se existir série) ou SS fixo
        ss_eff = ss_by_idx[i] if ss_by_idx is not None else safety_stock

        # Política de congelamento adotada: CONGELAR RECEBIMENTOS dentro do intervalo
        if i in frozen_idx:
            planned = 0
            # Apenas para relatório: quanto "faltaria" para atingir SS
            net = max(0, ss_eff - (on_hand_begin - gross))
            projected_end = on_hand_begin - gross + planned
        else:
            net = 0
            if on_hand_begin - gross < ss_eff:
                net = ss_eff - (on_hand_begin - gross)
                if lot_policy.upper() == "L4L":
                    planned = net
                else:
                    planned = int(math.ceil(net / lot_size) * lot_size)
            else:
                planned = 0
            projected_end = on_hand_begin - gross + planned

        por[i] = planned
        rows.append({
            "period": periods[i],
            "period_label": _label_pt(periods[i]),
            "gross_requirements": gross,
            "scheduled_receipts": rec_sched,
            "on_hand_begin": on_hand_begin,
            "net_requirements": max(0, net),
            "planned_order_receipts": planned,
            "projected_on_hand_end": projected_end,
        })
        on_hand = projected_end

    # liberações com lead time (em meses)
    for i in range(T):
        rel_idx = i - lead_time
        if rel_idx >= 0:
            pol[rel_idx] = por[i]

    df = pd.DataFrame(rows)
    df["planned_order_releases"] = pol

    # ATP por mês
    if cust is not None:
        df = _compute_atp_monthly(df, cust)

    cols = [
        "period", "period_label", "gross_requirements", "scheduled_receipts", "on_hand_begin",
        "net_requirements", "planned_order_receipts", "planned_order_releases", "projected_on_hand_end"
    ]
    if "atp" in df.columns:
        cols.append("atp")
    return df[cols]

def _compute_atp_monthly(df: pd.DataFrame, customer_orders: List[int]) -> pd.DataFrame:
    """ATP mensal com os mesmos princípios bucket-a-bucket."""
    T = len(df)
    rec = df["planned_order_receipts"].tolist()
    receive_idx = [i for i, x in enumerate(rec) if x > 0]
    atp = [0]*T

    def sum_orders(i, j):
        return sum(customer_orders[i:j+1]) if j >= i else 0

    if T > 0:
        first = 0
        next_rcv = next((k for k in receive_idx if k > first), T-1)
        supply = int(df.loc[first, "on_hand_begin"]) + int(df.loc[first, "planned_order_receipts"])
        demand = sum_orders(first, next_rcv)
        atp[first] = max(0, supply - demand)

    for idx in receive_idx:
        next_idx = next((k for k in receive_idx if k > idx), T-1)
        supply = int(df.loc[idx, "planned_order_receipts"])
        demand = sum_orders(idx, next_idx)
        atp[idx] = max(0, supply - demand)

    out = df.copy()
    out["atp"] = atp
    return out

# ======================= Exportação Excel (layout mensal) =======================

def export_mps_to_excel_monthly(
    df_mps: pd.DataFrame,
    filename: str,
    *,
    item_name: str = "Item",
    lot_size_info: int = 0,
    lead_time_info: int = 0,
    initial_inventory: int = 0,                # Em mão
    forecast_df: Optional[pd.DataFrame] = None,         # colunas ds,y (Previsto)
    orders_df: Optional[pd.DataFrame] = None,           # colunas ds,y (Em carteira)
    output_dir: Path = OUTPUT_DIR,                      # <-- SALVA AQUI
) -> str:
    """Cria planilha com colunas por mês (labels tipo 'Set/25') no caminho fixo."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # garante pasta
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / filename

    months = df_mps["period"].tolist()
    labels = df_mps["period_label"].tolist()
    T = len(labels)

    # Linhas
    if forecast_df is not None:
        fore = (forecast_df.assign(period=lambda d: d["ds"].map(_to_period_m))
                          .groupby("period")["y"].sum())
        previsto = [int(round(fore.get(m, 0))) for m in months]
    else:
        previsto = df_mps["gross_requirements"].astype(int).tolist()

    if orders_df is not None:
        od = (orders_df.assign(period=lambda d: d["ds"].map(_to_period_m))
                      .groupby("period")["y"].sum())
        em_carteira = [int(round(od.get(m, 0))) for m in months]
    else:
        em_carteira = [0]*T

    estoque_proj = df_mps["projected_on_hand_end"].astype(int).tolist()
    qtd_mps = df_mps["planned_order_receipts"].astype(int).tolist()
    inicio_mps = df_mps["planned_order_releases"].astype(int).tolist()
    atp_cum = df_mps.get("atp", pd.Series([0]*T)).astype(int).cumsum().tolist()

    # Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "MPS"

    bold = Font(bold=True)
    green = Font(color="006400")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="000000")
    border_all = Border(top=thin, bottom=thin, left=thin, right=thin)
    gray_fill = PatternFill("solid", fgColor="DDDDDD")

    start_col = 2  # B
    col_labels = 1
    row_item = 1
    row_emmao = 2
    row_mes_header = 3
    row_previsto = 4
    row_emcarteira = 5
    row_estoque = 6
    row_qtd_mps = 7
    row_inicio_mps = 8
    row_atp = 9

    last_col = start_col + T - 1

    # Cabeçalho
    ws.merge_cells(start_row=row_item, start_column=col_labels, end_row=row_item, end_column=last_col)
    ws.cell(row=row_item, column=col_labels).value = f"Item: {item_name}    Lote: {lot_size_info} unids.    Lead time: {lead_time_info}"
    ws.cell(row=row_item, column=col_labels).font = bold
    ws.cell(row=row_item, column=col_labels).alignment = left

    ws.merge_cells(start_row=row_emmao, start_column=col_labels, end_row=row_emmao, end_column=col_labels+1)
    ws.cell(row=row_emmao, column=col_labels).value = f"Em mão: {int(initial_inventory)}"
    ws.cell(row=row_emmao, column=col_labels).alignment = left
    ws.cell(row=row_emmao, column=col_labels).fill = gray_fill
    ws.cell(row=row_emmao, column=col_labels).border = border_all

    ws.cell(row=row_mes_header, column=col_labels).value = "mês"
    ws.cell(row=row_mes_header, column=col_labels).font = bold
    ws.cell(row=row_mes_header, column=col_labels).alignment = center

    for j, lab in enumerate(labels, start=start_col):
        c = ws.cell(row=row_mes_header, column=j, value=lab)
        c.alignment = center
        c.font = bold

    # Rótulos
    ws.cell(row=row_previsto, column=col_labels).value = "Previsto"
    ws.cell(row=row_emcarteira, column=col_labels).value = "Em carteira"
    ws.cell(row=row_estoque, column=col_labels).value = "Estoque Proj."
    ws.cell(row=row_qtd_mps, column=col_labels).value = "Qtde. MPS"
    ws.cell(row=row_inicio_mps, column=col_labels).value = "Início MPS"
    ws.cell(row=row_atp, column=col_labels).value = "ATP(cum)"
    for r in [row_previsto, row_emcarteira, row_estoque, row_qtd_mps, row_inicio_mps, row_atp]:
        ws.cell(row=r, column=col_labels).font = bold
        ws.cell(row=r, column=col_labels).alignment = left

    # Escreve linhas
    def write_row(values, row_idx, font=None):
        for j, v in enumerate(values, start=start_col):
            c = ws.cell(row=row_idx, column=j, value=int(v))
            c.alignment = center
            if font:
                c.font = font

    write_row(previsto, row_previsto, font=green)
    write_row(em_carteira, row_emcarteira)
    write_row(estoque_proj, row_estoque)
    write_row(qtd_mps, row_qtd_mps)
    write_row(inicio_mps, row_inicio_mps)
    write_row(atp_cum, row_atp)

    # Bordas / largura
    max_row = row_atp
    for r in range(row_mes_header, max_row + 1):
        for c in range(col_labels, last_col + 1):
            ws.cell(row=r, column=c).border = border_all

    ws.column_dimensions["A"].width = 16
    for c in range(start_col, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 10

    ws.freeze_panes = "B4"

    wb.save(str(out_path))
    return str(out_path)

# ======================= Demonstração rápida =======================
if __name__ == "__main__":
    # Exemplo simples
    df_forecast = pd.DataFrame({
        "ds": ["Set/25", "Out/25", "Nov/25", "Dez/25", "Jan/26", "Fev/26"],
        "y":  [300,      300,      300,      300,      350,      350     ]
    })

    df_orders = pd.DataFrame({
        "ds": ["Set/25", "Out/25", "Nov/25", "Dez/25"],
        "y":  [280,       120,       40,        0]
    })

    # SS variável (apenas para demo): 6 valores, um por mês
    ss_series_demo = [50, 70, 50, 60, 80, 80]

    params = dict(
        lot_policy="FX",
        lot_size=150,
        safety_stock=0,
        lead_time=1,             # 1 mês
        initial_inventory=55,
        scheduled_receipts={},   # ex.: {"Jan/26": 50}
        safety_stock_series=ss_series_demo,
        frozen_range=("Out/25", "Nov/25"),   # exemplo de fence
    )

    mps = compute_mps_monthly(
        df_forecast,
        **params,
        firm_customer_orders=df_orders,
    )

    out = export_mps_to_excel_monthly(
        mps, filename="MPS_mensal.xlsx",
        item_name="Cadeira de ripas",
        lot_size_info=params["lot_size"],
        lead_time_info=params["lead_time"],
        initial_inventory=params["initial_inventory"],
        forecast_df=df_forecast,
        orders_df=df_orders,
    )
    print(f"Gerado em: {out}")
